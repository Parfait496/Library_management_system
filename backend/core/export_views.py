from django.http import HttpResponse
from django.utils import timezone
from rest_framework import permissions
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.contrib.auth import get_user_model

from books.models import Book, Genre
from borrowing.models import BorrowRecord
from fines.models import Fine

User = get_user_model()


class IsAdminUser(permissions.BasePermission):
    def has_permission(self, request, view):
        return (
            request.user.is_authenticated and
            request.user.role == 'ADMIN'
        )


def _write_sheet(ws, headers, rows):
    """Writes a header row (styled) + data rows, then auto-fits columns."""
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='16333A', end_color='16333A', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(length + 2, 40)


class DatabaseSnapshotExportAPIView(APIView):
    """
    GET /api/export/snapshot/
    GET /api/export/snapshot/?sheets=books
    GET /api/export/snapshot/?sheets=members
    GET /api/export/snapshot/?sheets=books,genres

    Admin-only. Generates an Excel snapshot. With no ?sheets param,
    includes everything (Books, Genres, Borrow Records, Fines,
    Members). Pass a comma-separated ?sheets= to export just one or
    a few — e.g. the Books page's "Export" button calls this with
    ?sheets=books to get a single-sheet Books-only workbook.
    """
    permission_classes = [IsAdminUser]

    def get(self, request):
        requested = request.query_params.get('sheets')
        wanted = (
            {s.strip().lower() for s in requested.split(',')}
            if requested else
            {'books', 'genres', 'borrowing', 'fines', 'members'}
        )

        wb = Workbook()
        first_sheet_used = False

        def get_ws(title):
            nonlocal first_sheet_used
            if not first_sheet_used:
                ws = wb.active
                ws.title = title
                first_sheet_used = True
                return ws
            return wb.create_sheet(title)

        # ---- Books ----
        if 'books' in wanted:
            ws_books = get_ws('Books')
            _write_sheet(
                ws_books,
                ['ID', 'Title', 'Author', 'ISBN', 'Publisher', 'Year',
                 'Genre', 'Category Path', 'Total Copies', 'Available Copies',
                 'Added By', 'Created At'],
                [
                    [
                        b.id, b.title, b.author, b.isbn or '', b.publisher or '',
                        b.publication_year or '',
                        b.genre.name if b.genre else '',
                        b.genre.full_path if b.genre else '',
                        b.total_copies, b.available_copies,
                        b.added_by.username if b.added_by else '',
                        b.created_at.strftime('%Y-%m-%d %H:%M'),
                    ]
                    for b in Book.objects.select_related('genre', 'added_by').all()
                ]
            )

        # ---- Genres ----
        if 'genres' in wanted:
            ws_genres = get_ws('Genres')
            _write_sheet(
                ws_genres,
                ['ID', 'Name', 'Parent Category', 'Full Path', 'Description'],
                [
                    [g.id, g.name, g.parent.name if g.parent else '', g.full_path, g.description or '']
                    for g in Genre.objects.select_related('parent').all()
                ]
            )

        # ---- Borrow Records ----
        if 'borrowing' in wanted:
            ws_borrow = get_ws('Borrow Records')
            _write_sheet(
                ws_borrow,
                ['ID', 'Member', 'Book', 'Status', 'Request Date', 'Due Date',
                 'Return Date', 'Days Overdue', 'Fine Amount (RWF)', 'Processed By'],
                [
                    [
                        r.id, r.member.username, r.book.title, r.status,
                        r.request_date.strftime('%Y-%m-%d'),
                        r.due_date.strftime('%Y-%m-%d') if r.due_date else '',
                        r.return_date.strftime('%Y-%m-%d') if r.return_date else '',
                        r.days_overdue, r.fine_amount,
                        r.processed_by.username if r.processed_by else '',
                    ]
                    for r in BorrowRecord.objects.select_related('member', 'book', 'processed_by').all()
                ]
            )

        # ---- Fines ----
        if 'fines' in wanted:
            ws_fines = get_ws('Fines')
            _write_sheet(
                ws_fines,
                ['ID', 'Member', 'Book', 'Amount (RWF)', 'Days Overdue', 'Status',
                 'Issued Date', 'Resolved Date', 'Resolved By', 'Note'],
                [
                    [
                        f.id, f.member.username, f.borrow_record.book.title,
                        float(f.amount), f.days_overdue, f.status,
                        f.issued_date.strftime('%Y-%m-%d'),
                        f.resolved_date.strftime('%Y-%m-%d') if f.resolved_date else '',
                        f.resolved_by.username if f.resolved_by else '',
                        f.note or '',
                    ]
                    for f in Fine.objects.select_related('member', 'borrow_record__book', 'resolved_by').all()
                ]
            )

        # ---- Members ----
        if 'members' in wanted:
            ws_members = get_ws('Members')
            _write_sheet(
                ws_members,
                ['ID', 'Username', 'Full Name', 'Email', 'Role', 'Phone', 'Student ID', 'Joined'],
                [
                    [
                        u.id, u.username, u.full_name, u.email, u.role,
                        u.phone_number or '', u.student_id or '',
                        u.created_at.strftime('%Y-%m-%d'),
                    ]
                    for u in User.objects.all()
                ]
            )

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        label = requested.replace(',', '_') if requested else 'full'
        filename = f'asome_library_{label}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
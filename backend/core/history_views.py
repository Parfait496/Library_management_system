from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from books.models import Book, Genre
from borrowing.models import BorrowRecord
from fines.models import Fine

# NOTE: this endpoint depends on django-simple-history being installed
# and HistoricalRecords() added to each model — see
# audit-trail-setup.md. Until that's done, this returns an empty list
# for any model that doesn't have `.history` yet (it won't error).

ACTION_LABELS = {'+': 'Created', '~': 'Updated', '-': 'Deleted'}


class IsStaffUser(permissions.BasePermission):
    def has_permission(self, request, view):
        return (
            request.user.is_authenticated and
            request.user.role in ['ADMIN', 'LIBRARIAN']
        )


def _describe_book(record):
    if record.history_type == '+':
        return f'"{record.title}" added to the catalogue ({record.total_copies} copies)'
    if record.history_type == '-':
        return f'"{record.title}" removed from the catalogue'
    # Updated — show exactly what changed, not just "updated"
    prev = record.prev_record
    if prev:
        delta = record.diff_against(prev)
        if delta.changes:
            changed = ', '.join(
                f'{c.field} {c.old} → {c.new}' for c in delta.changes
            )
            return f'"{record.title}" updated ({changed})'
    return f'"{record.title}" updated'


def _describe_genre(record):
    parent_note = ''
    if record.parent_id:
        parent = Genre.objects.filter(pk=record.parent_id).first()
        if parent:
            parent_note = f' (under {parent.name})'
    if record.history_type == '+':
        return f'"{record.name}"{parent_note} created'
    if record.history_type == '-':
        return f'"{record.name}"{parent_note} deleted'
    return f'"{record.name}"{parent_note} updated'


def _describe_borrow(record):
    book_title = record.book.title if record.book_id else 'a book'
    member_name = record.member.username if record.member_id else 'a member'
    action_by_status = {
        'REQUESTED': f'"{book_title}" requested by {member_name}',
        'APPROVED':  f'request for "{book_title}" approved ({member_name})',
        'REJECTED':  f'request for "{book_title}" rejected ({member_name})',
        'BORROWED':  f'"{book_title}" marked as borrowed by {member_name}',
        'RETURNED':  f'"{book_title}" marked as returned by {member_name}',
        'OVERDUE':   f'"{book_title}" flagged overdue ({member_name})',
    }
    return action_by_status.get(record.status, f'"{book_title}" record updated ({member_name})')


def _describe_fine(record):
    member_name = record.member.username if record.member_id else 'a member'
    if record.history_type == '+':
        return f'{record.amount} RWF fine issued to {member_name} ({record.days_overdue} days overdue)'
    if record.status == 'PAID':
        return f'{member_name}\u2019s {record.amount} RWF fine marked as paid'
    if record.status == 'WAIVED':
        return f'{member_name}\u2019s {record.amount} RWF fine waived'
    return f'fine for {member_name} updated'


DESCRIBERS = {
    'Book':          _describe_book,
    'Genre':         _describe_genre,
    'Borrow Record': _describe_borrow,
    'Fine':          _describe_fine,
}


def _model_history(model_cls, label):
    if not hasattr(model_cls, 'history'):
        return []
    describe = DESCRIBERS[label]
    entries = []
    for record in model_cls.history.all().select_related('history_user')[:200]:
        entries.append({
            'model': label,
            'object_id': record.id,
            'description': describe(record),
            'action': ACTION_LABELS.get(record.history_type, record.history_type),
            'changed_by': record.history_user.username if record.history_user else 'System',
            'date': record.history_date.isoformat(),
        })
    return entries


class ActivityLogAPIView(APIView):
    """
    GET /api/activity-log/

    Returns the last 100 changes across Books, Genres, Borrow Records,
    and Fines, newest first, in plain-English form — no Django admin
    or shell access needed to see "who changed what, when."
    """
    permission_classes = [IsStaffUser]

    def get(self, request):
        entries = []
        entries += _model_history(Book, 'Book')
        entries += _model_history(Genre, 'Genre')
        entries += _model_history(BorrowRecord, 'Borrow Record')
        entries += _model_history(Fine, 'Fine')

        entries.sort(key=lambda e: e['date'], reverse=True)
        return Response(entries[:100])


class ActivityLogExportAPIView(APIView):
    """GET /api/export/activity-log/ — same data as ActivityLogAPIView, as .xlsx"""
    permission_classes = [IsStaffUser]

    def get(self, request):
        entries = []
        entries += _model_history(Book, 'Book')
        entries += _model_history(Genre, 'Genre')
        entries += _model_history(BorrowRecord, 'Borrow Record')
        entries += _model_history(Fine, 'Fine')
        entries.sort(key=lambda e: e['date'], reverse=True)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Activity Log'
        headers = ['Date', 'Model', 'Action', 'Description', 'Changed By']
        ws.append(headers)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='16333A', end_color='16333A', fill_type='solid')
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for e in entries:
            ws.append([e['date'], e['model'], e['action'], e['description'], e['changed_by']])

        for col in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(length + 2, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'asome_library_activity_log_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response